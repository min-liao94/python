import os
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook


def main():
    # 以腳本所在位置為基準
    base_dir = os.path.dirname(os.path.abspath(__file__))
    data_dir = os.path.join(base_dir, "data")
    list_file = os.path.join(base_dir, "list_all.xlsx")
    log_file = os.path.join(base_dir, "completeness_log.csv")

    if not os.path.isdir(data_dir):
        print(f"找不到資料夾：{data_dir}")
        return

    if not os.path.exists(list_file):
        print(f"找不到清單檔案：{list_file}")
        return

    # 讀取 list_all.xlsx
    df = pd.read_excel(list_file)
    if df.shape[1] < 1:
        print("list_all.xlsx 至少需要第一欄（工號）。")
        return

    total = 0       # 總問卷數（以清單為基準）
    finished = 0    # 已完成（F1 = 'ok'）
    status_list = []  # 對應每一列要寫回去的 V / X
    update_time_list = []  # 對應每一列要寫回去的更新時間

    # 檢查是否已有「完成狀態」和「更新時間」欄位
    has_status_col = "完成狀態" in df.columns
    has_time_col = "更新時間" in df.columns

    # 欄位假設：第 0 欄為工號，第 1 欄（若存在）為姓名
    for idx, row in df.iterrows():
        gonghao_raw = row.iloc[0]
        name_raw = row.iloc[1] if df.shape[1] > 1 else ""

        if pd.isna(gonghao_raw):
            status_list.append("")
            update_time_list.append("")
            continue

        gonghao = str(gonghao_raw).strip()
        name = str(name_raw).strip() if pd.notna(name_raw) else ""

        if not gonghao or gonghao.lower() in ["nan", "none", ""]:
            status_list.append("")
            update_time_list.append("")
            continue

        # 取得原本的完成狀態（如果有的話）
        old_status = ""
        if has_status_col:
            old_status_cell = row.get("完成狀態", "")
            if pd.notna(old_status_cell):
                old_status = str(old_status_cell).strip()

        # 取得原本的更新時間（如果有的話）
        old_update_time = ""
        if has_time_col:
            old_time_cell = row.get("更新時間", "")
            if pd.notna(old_time_cell):
                old_update_time = str(old_time_cell).strip()

        total += 1

        # 對應 distribute.py 建立檔案的路徑與檔名
        gonghao_folder = os.path.join(data_dir, gonghao)
        files_folder = os.path.join(gonghao_folder, "files")
        dest_filename = (
            f"2025資訊發展中心問卷_{name}.xlsx"
            if name
            else "2025資訊發展中心問卷.xlsx"
        )
        file_path = os.path.join(files_folder, dest_filename)

        done = False
        file_update_time = ""
        if os.path.exists(file_path):
            try:
                wb = load_workbook(file_path, data_only=True)
                ws = wb.active
                value = ws["F1"].value
                if isinstance(value, str) and value.strip().lower() == "ok":
                    done = True
                
                # 取得檔案更新時間
                try:
                    mtime = os.path.getmtime(file_path)
                    file_update_time = datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M:%S")
                except Exception:
                    file_update_time = ""
            except Exception as e:
                print(f"讀取檔案發生錯誤：{file_path} ({e})")
        else:
            print(f"找不到問卷檔案：{file_path}")

        if done:
            finished += 1
            status_list.append("V")
            # 只有當完成狀態從非 V 變成 V 時，才更新時間
            if old_status != "V":
                # 狀態剛變成 V，更新時間
                update_time_list.append(file_update_time)
            else:
                # 原本就是 V，保留原有時間
                update_time_list.append(old_update_time)
        else:
            status_list.append("X")
            # 未完成時，保留原有時間（如果有的話）
            update_time_list.append(old_update_time)

    if total == 0:
        print("清單中沒有有效的工號資料。")
        return

    rate = finished / total * 100
    print(f"完成 {finished}/{total}，完成度：{rate:.2f}%")

    # 把 V / X 和更新時間寫回 list_all.xlsx
    df["完成狀態"] = status_list
    df["更新時間"] = update_time_list
    df.to_excel(list_file, index=False)
    print(f"已將完成狀態（V/X）和更新時間寫回：{list_file}")

    # 每次執行都把結果記錄到同目錄的 completeness_log.csv
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"{now},{finished},{total},{rate:.2f}\n"

    if not os.path.exists(log_file):
        with open(log_file, "w", encoding="utf-8") as f:
            f.write("datetime,finished,total,rate_percent\n")
            f.write(line)
    else:
        with open(log_file, "a", encoding="utf-8") as f:
            f.write(line)


if __name__ == "__main__":
    main()
