import os
import pandas as pd
import shutil
from openpyxl import load_workbook

# 取得當前腳本所在目錄
script_dir = os.path.dirname(os.path.abspath(__file__))
os.chdir(script_dir)

# 讀取 list_c.xlsx 檔案
list_file = "list_all.xlsx"
questionnaire_file = "資訊發展中心問卷.xlsx"

print(f"當前工作目錄：{os.getcwd()}")
print(f"讀取檔案：{list_file}")

# 讀取 Excel 檔案
try:
    df = pd.read_excel(list_file)

except Exception as e:
    print(f"讀取檔案錯誤：{e}")
    exit(1)


if len(df.columns) >= 2:
    # 取得第一欄和第二欄的資料
    first_col = df.iloc[:, 0]  # 第一欄（索引0）
    second_col = df.iloc[:, 1]  # 第二欄（索引1）
    
    print(f"\n第一欄名稱：{df.columns[0]}")
    print(f"第二欄名稱：{df.columns[1]}")
    
    # 建立 data 資料夾
    data_folder = "data"
    if not os.path.exists(data_folder):
        os.makedirs(data_folder)
        print(f"\n已建立資料夾：{data_folder}")
    else:
        print(f"\n資料夾已存在：{data_folder}")
    
    # 檢查問卷檔案是否存在
    if not os.path.exists(questionnaire_file):
        print(f"錯誤：找不到檔案 {questionnaire_file}")
        exit(1)
    
    # 計數器
    processed_count = 0
    
    # 遍歷每一行資料
    for idx in range(len(df)):
        gonghao_raw = first_col.iloc[idx]
        name_raw = second_col.iloc[idx]
        
        # 處理工號
        if pd.isna(gonghao_raw):
            continue
        gonghao = str(gonghao_raw).strip()
        if not gonghao or gonghao.lower() in ['nan', 'none', '']:
            continue
        
        # 處理姓名
        name = str(name_raw).strip() if pd.notna(name_raw) else ""
        if name.lower() in ['nan', 'none']:
            name = ""
        
        # 建立工號資料夾
        gonghao_folder = os.path.join(data_folder, gonghao)
        files_folder = os.path.join(gonghao_folder, "files")
        if not os.path.exists(files_folder):
            os.makedirs(files_folder)
            print(f"已建立工號資料夾：{files_folder}")
        
        # 複製問卷檔案到工號資料夾，並將姓名寫入 E1 欄位
        # 檔名改為「2025資訊發展中心問卷_姓名.xlsx」
        dest_filename = f"2025資訊發展中心問卷_{name}.xlsx" if name else "2025資訊發展中心問卷.xlsx"
        dest_file = os.path.join(files_folder, dest_filename)
        if not os.path.exists(dest_file):
            # 先複製原始檔案
            shutil.copy2(questionnaire_file, dest_file)
            
            # 讀取問卷檔案並修改 E1 欄位
            try:
                wb = load_workbook(dest_file)
                ws = wb.active
                # 將姓名寫入 E1 欄位
                ws['E1'] = name
                wb.save(dest_file)
                print(f"已複製問卷並寫入姓名到：{dest_file} (工號：{gonghao}, 姓名：{name})")
                processed_count += 1
            except Exception as e:
                print(f"寫入姓名時發生錯誤：{e} (工號：{gonghao}, 姓名：{name})")
        else:
            print(f"檔案已存在，跳過：{dest_file} (工號：{gonghao}, 姓名：{name})")
    
    print(f"\n處理完成！共處理 {processed_count} 筆資料")
else:
    print("錯誤：Excel 檔案至少需要兩欄資料")

