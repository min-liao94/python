import os
from pathlib import Path
from typing import Dict, List, Any

import pandas as pd
from openpyxl import load_workbook


def extract_from_file(
    gonghao: str,
    file_path: Path,
    max_questions_ref: List[int],
    question_texts_ref: List[str],
) -> List[Dict[str, Any]]:
    """
    從單一問卷檔案中：
    - 檢查 F1 是否為 'ok'
    - 讀取 E1（填答者姓名）
    - 從表格中整理出「受測者 / 分數 / 建議」
    """
    result: List[Dict[str, Any]] = []

    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
    except Exception as e:
        print(f"讀取檔案失敗：{file_path}（{e}）")
        return result

    # 1. 確認 F1 = ok
    f1 = ws["F1"].value
    if not (isinstance(f1, str) and f1.strip().lower() == "ok"):
        # 未完成問卷，直接略過
        return result

    # 2. 取得填答者姓名（E1）
    name_cell = ws["E1"].value
    if isinstance(name_cell, str):
        name = name_cell.strip()
    elif name_cell is not None:
        name = str(name_cell).strip()
    else:
        name = ""

    # 3. 題目與受測者結構：
    #    - 題目文字：放在 D4～D11（縱向，一列一題）
    #    - 各受測者：放在某一列（例如第 3 列）從 E 欄開始，每欄一位受測者
    #    - 分數：在題目列與受測者欄的交會處

    # 3-1. 先抓題目列（D4～往下，直到遇到空白或超過 D11）
    question_rows: List[int] = []
    row_idx = 4
    while row_idx <= 11:  # 依需求限制在 D4～D11
        cell = ws.cell(row=row_idx, column=4)  # D 欄
        cell_val = cell.value
        if isinstance(cell_val, str):
            cell_val = cell_val.strip()
        if not cell_val:
            break
        question_rows.append(row_idx)
        # 同步記錄題目文字，用於輸出表頭（詳細問題）
        title = str(cell_val)
        idx = len(question_rows) - 1  # 0-based
        if idx >= len(question_texts_ref):
            question_texts_ref.append(title)
        elif not question_texts_ref[idx]:
            question_texts_ref[idx] = title
        row_idx += 1

    if not question_rows:
        print(f"找不到 D4～D11 之間的題目文字：{file_path}")
        return result

    # 記錄全域的最大題數（用在產出欄位 問1、問2...）
    if max_questions_ref[0] < len(question_rows):
        max_questions_ref[0] = len(question_rows)

    # 3-2. 受測者名稱：假設在第 3 列，從 E 欄開始往右
    subject_header_row = 3
    subject_cols: Dict[int, str] = {}
    for col_idx in range(5, ws.max_column + 1):  # E 欄開始
        cell_val = ws.cell(row=subject_header_row, column=col_idx).value
        if isinstance(cell_val, str):
            cell_val = cell_val.strip()
        if not cell_val:
            continue
        subject_cols[col_idx] = str(cell_val)

    if not subject_cols:
        print(f"找不到第 {subject_header_row} 列 E 欄之後的受測者名稱：{file_path}")
        return result

    num_questions = len(question_rows)

    # 4. 對每一位受測者彙整「問1～問N」分數，一列一位受測者
    for col_idx, subject in subject_cols.items():
        scores: Dict[str, Any] = {}
        has_score = False

        for q_idx, q_row in enumerate(question_rows, start=1):
            value = ws.cell(row=q_row, column=col_idx).value
            if isinstance(value, str):
                value = value.strip()
            if value not in (None, ""):
                has_score = True

            # 使用實際題目文字當作 key，確保和輸出欄位名稱一致
            q_index = q_idx - 1
            if q_index < len(question_texts_ref) and question_texts_ref[q_index]:
                col_name = question_texts_ref[q_index]
            else:
                col_name = f"問{q_idx}"

            if value in (None, ""):
                # 依題目位置補預設：最後兩題補 None，其餘補 0
                if q_index >= num_questions - 2:
                    filled_value = "None"
                else:
                    filled_value = 0
            else:
                filled_value = value

            scores[col_name] = filled_value

        if not has_score:
            # 此受測者對所有題目都沒填，略過
            continue

        record: Dict[str, Any] = {"工號": gonghao, "姓名": name, "受測者": subject}
        record.update(scores)
        result.append(record)

    return result


def main() -> None:
    base_dir = Path(os.path.dirname(os.path.abspath(__file__)))
    data_dir = base_dir / "data"

    if not data_dir.is_dir():
        print(f"找不到 data 資料夾：{data_dir}")
        return

    all_records: List[Dict[str, Any]] = []
    # 用一個 list 包住整數，讓 extract_from_file 可以更新題目總數
    max_questions_ref: List[int] = [0]
    # 紀錄實際題目文字，之後拿來當欄位名稱（詳細問題）
    question_texts_ref: List[str] = []

    # data 目錄底下：每一個工號一個資料夾
    for gonghao_dir in sorted(data_dir.iterdir()):
        if not gonghao_dir.is_dir():
            continue

        gonghao = gonghao_dir.name
        files_dir = gonghao_dir / "files"
        if not files_dir.is_dir():
            continue

        # 可能檔名為「2025資訊發展中心問卷_姓名.xlsx」或「2025資訊發展中心問卷.xlsx」
        for file_path in files_dir.glob("2025資訊發展中心問卷*.xlsx"):
            print(f"處理檔案：{file_path}")
            records = extract_from_file(
                gonghao, file_path, max_questions_ref, question_texts_ref
            )
            all_records.extend(records)

    if not all_records:
        print("沒有任何符合條件（F1 = ok）的問卷資料。")
        return

    # 欄位順序：工號、姓名、受測者、更新時間、各題詳細問題
    num_q = max_questions_ref[0]
 
    question_columns: List[str] = []
    for i in range(num_q):
        if i < len(question_texts_ref) and question_texts_ref[i]:
            question_columns.append(question_texts_ref[i])
        else:
            question_columns.append(f"問{i + 1}")
    columns = ["工號", "姓名", "受測者"] + question_columns
    df = pd.DataFrame(all_records, columns=columns)
    output_file = base_dir / "2025資訊發展中心問卷_彙整結果.xlsx"
    df.to_excel(output_file, index=False)
    print(f"已輸出彙整結果至：{output_file}")


if __name__ == "__main__":
    main()


